from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from typing import List, Tuple, Optional
from .category_mapper import CategoryMapper
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from .chart_generator import ChartGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generates formatted Excel financial statements."""
    
    def __init__(self, transactions_df: pd.DataFrame, industry: str):
        """
        Initialize Excel generator with transaction data.
        
        Args:
            transactions_df: DataFrame containing categorized transactions
            industry: Industry type for category mapping
        """
        self.transactions_df = transactions_df.copy()
        self.category_mapper = CategoryMapper(industry)
        self.wb = Workbook()
        
        # Excel styles
        self.header_font = Font(bold=True)
        self.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin')
        )
        self.total_border = Border(
            bottom=Side(style='double'),
            top=Side(style='thin')
        )
        
    def generate(self, company_name: str) -> str:
        """
        Generate Excel file with Transactions and Statements sheets.
        
        Args:
            company_name: Name of company for filename
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Setup sheets
            self._setup_transactions_sheet()
            self._setup_statements_sheet()
            
            # Generate filename and save
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{company_name}_{timestamp}.xlsx"
            output_dir = Path(__file__).parent.parent.parent / 'statements_output'
            output_dir.mkdir(exist_ok=True)
            output_path = output_dir / filename
            
            # Save workbook
            self.wb.save(str(output_path))
            
            logger.info(f"Generated financial statements: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {str(e)}")
            raise
        finally:
            self.wb.close()

    def _setup_transactions_sheet(self):
        """Create and format Transactions sheet."""
        # Sort transactions by date
        self.transactions_df.sort_values('Date', inplace=True)
        
        # Get or create Transactions sheet
        ws = self.wb.active
        ws.title = 'Transactions'
        
        # Write headers
        headers = ['Date', 'End of Month', 'Description', 'Category', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.border = self.border
        
        # Write data
        for row_idx, row in enumerate(self.transactions_df.itertuples(), 2):
            # Date column
            ws.cell(row=row_idx, column=1, value=row.Date)
            ws.cell(row=row_idx, column=1).number_format = 'mm/dd/yyyy'
            
            # End of Month column - use EOMONTH formula
            end_month_cell = ws.cell(row=row_idx, column=2)
            end_month_cell.value = f'=EOMONTH(A{row_idx},0)'
            end_month_cell.number_format = 'mm/dd/yyyy'
            
            ws.cell(row=row_idx, column=3, value=row.Description)
            ws.cell(row=row_idx, column=4, value=row.Category)
            amount_cell = ws.cell(row=row_idx, column=5, value=row.Amount)
            amount_cell.number_format = '#,##0.00_);(#,##0.00)'
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _get_date_range(self) -> List[datetime]:
        """
        Get list of all months between earliest and latest transaction.
        
        Returns:
            List of datetime objects for each month
        """
        dates = pd.to_datetime(self.transactions_df['End_of_Month'])
        start_date = dates.min()
        end_date = dates.max()
        
        # Generate all months between start and end
        months = pd.date_range(
            start=start_date,
            end=end_date,
            freq='ME'
        )
        
        return months.tolist()

    def _setup_statements_sheet(self):
        """Create and format Statements sheet."""
        ws = self.wb.create_sheet('Statements')
        
        # Hide column A and set white background
        ws.column_dimensions['A'].hidden = True
        ws.sheet_view.showGridLines = False
        
        # Get category structure and date range
        categories = self.category_mapper.get_category_structure()
        months = self._get_date_range()
        
        # Write month headers with full date and custom format
        for col, month in enumerate(months, 3):
            cell = ws.cell(row=1, column=col)
            cell.value = month.replace(hour=0, minute=0, second=0, microsecond=0)
            cell.number_format = 'mmm-yy'
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='right')
            cell.border = Border(bottom=Side(style='thin'))
        
        # Write categories and formulas
        current_row = 2  # Start at row 2 to match desired layout
        revenue_total_row = None
        cogs_total_row = None
        
        for category in categories:
            current_row += 1
            
            # Write category name
            ws.cell(row=current_row, column=1, value=category['prefix'])
            name_cell = ws.cell(row=current_row, column=2, value=category['name'])
            
            # Handle formatting
            if category['is_parent']:
                name_cell.font = self.header_font
            elif not category['is_parent'] and not category['is_total']:
                # Add indentation for subcategories
                name_cell.alignment = Alignment(indent=1)
            
            if category['is_total']:
                name_cell.font = self.header_font
                name_cell.alignment = Alignment(indent=1)  # Indent total rows
            
            # Track total rows
            if category['name'] == 'Total Revenue':
                revenue_total_row = current_row
            
            # Add formulas for each month
            for col, month in enumerate(months, 3):
                cell = ws.cell(row=current_row, column=col)
                
                if category['is_total']:
                    # Add total formula with only top border
                    start_row = current_row - len(self.category_mapper.get_subcategories(category['parent']))
                    formula = f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{current_row-1})"
                    cell.value = formula
                    cell.font = self.header_font
                    cell.border = Border(top=Side(style='thin'))  # Only top border for total rows
                elif not category['is_parent']:
                    # Add SUMIFS formula with negative sign for costs
                    prefix = "-" if ("Cost of Sales:" in category['prefix'] or "Operating Expenses:" in category['prefix']) else ""
                    formula = (
                        f'={prefix}SUMIFS(Transactions!$E:$E,'
                        f'Transactions!$D:$D,Statements!$A{current_row},'
                        f'Transactions!$B:$B,Statements!{get_column_letter(col)}$1)'
                    )
                    cell.value = formula
                
                cell.number_format = '#,##0_);(#,##0)'
                cell.alignment = Alignment(horizontal='right')
            
            # Track Cost of Sales total row and add Gross Income
            if category['name'] == 'Total Cost of Sales':
                cogs_total_row = current_row
                
                # Add Gross Income after Cost of Sales
                current_row += 2  # Add two rows for spacing
                ws.cell(row=current_row, column=1, value="Gross Income")
                gross_cell = ws.cell(row=current_row, column=2, value="Gross Income")
                gross_cell.font = self.header_font
                
                # Add Gross Income formulas
                for col, _ in enumerate(months, 3):
                    cell = ws.cell(row=current_row, column=col)
                    formula = f"={get_column_letter(col)}{revenue_total_row}-{get_column_letter(col)}{cogs_total_row}"
                    cell.value = formula
                    cell.number_format = '#,##0_);(#,##0)'
                    cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal='right')
                
                gross_income_row = current_row
                current_row += 1  # Add one row after Gross Income
            
            # Add extra blank row after other totals
            elif category['is_total'] and category['name'] != 'Total Cost of Sales':
                current_row += 1
        
        # Update EBITDA calculation to use Gross Income
        self._add_ebitda_calculation(ws, current_row + 1, len(months), gross_income_row)

        return ws

    def _add_ebitda_calculation(self, ws, row: int, num_months: int, gross_income_row: int):
        """Add EBITDA calculation row."""
        ws.cell(row=row, column=1, value="EBITDA")
        ebitda_cell = ws.cell(row=row, column=2, value="EBITDA")
        ebitda_cell.font = self.header_font
        
        # Find Operating Expenses total row
        opex_total_row = None
        for row_idx in range(1, row):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value == "Total Operating Expenses":
                opex_total_row = row_idx
                break
        
        # Add EBITDA formula for each month
        for col in range(3, num_months + 3):
            col_letter = get_column_letter(col)
            formula = f"={col_letter}{gross_income_row}-{col_letter}{opex_total_row}"
            
            cell = ws.cell(row=row, column=col)
            cell.value = formula
            cell.number_format = '#,##0_);(#,##0)'
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='right')