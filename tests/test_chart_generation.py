import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import logging

# Add root directory to Python path
root_dir = Path(__file__).parent.parent
sys.path.append(str(root_dir))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_basic_chart():
    """Test creating a single basic chart."""
    try:
        # Create a simple workbook with test data
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add some test data
        for row in range(1, 5):
            for col in range(1, 5):
                ws.cell(row=row, column=col, value=row*col)
        
        # Create Charts sheet
        charts_ws = wb.create_sheet('Charts')
        charts_ws.sheet_view.showGridLines = False
        
        # Create simple bar chart
        chart = BarChart()
        chart.title = "Test Chart"
        chart.height = 15
        chart.width = 30
        
        data = Reference(ws, min_col=1, min_row=1, max_col=4, max_row=4)
        chart.add_data(data)
        
        # Add chart to worksheet
        charts_ws.add_chart(chart, "B2")
        
        # Save workbook
        output_dir = root_dir / 'test_output'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / 'test_chart.xlsx'
        
        wb.save(str(output_path))
        logger.info(f"Test chart saved to: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error in test_basic_chart: {str(e)}")
        return False
    finally:
        wb.close()

def main():
    """Run chart generation tests."""
    logger.info("Testing basic chart generation...")
    if test_basic_chart():
        logger.info("✓ Basic chart test passed")
    else:
        logger.error("✗ Basic chart test failed")

if __name__ == "__main__":
    main() 