import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
import logging

# Add root directory to Python path
root_dir = Path(__file__).parent.parent
sys.path.append(str(root_dir))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_stacked_bar_only():
    """Test creating just a stacked bar chart without the line."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add test data - first row is months
        months = ['Jan-24', 'Feb-24']
        ws.cell(row=1, column=1, value='Category')  # Header for category column
        for col, month in enumerate(months, 2):
            ws.cell(row=1, column=col, value=month)
        
        # Add category data - each row is a category's values across months
        data = [
            ('Cat A', 100, 50),   # Values for Cat A
            ('Cat B', 120, 70),   # Values for Cat B
        ]
        
        for row, (category, *values) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=category)
            for col, value in enumerate(values, 2):
                ws.cell(row=row, column=col, value=value)
        
        # Log the data for debugging
        logger.info("Worksheet data:")
        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
            logger.info(row)
        
        # Create Charts sheet
        charts_ws = wb.create_sheet('Charts')
        charts_ws.sheet_view.showGridLines = False
        
        # Create stacked bar chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Test Stacked Chart"
        chart.height = 15
        chart.width = 30
        
        # Categories (X axis) - the months
        cats = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1)
        chart.set_categories(cats)
        
        # Data - all category values
        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=False)
        
        # Set series names from first column
        series_names = Reference(ws, min_col=1, min_row=2, max_row=3)
        chart.set_categories(cats)
        chart.series[0].title = 'Cat A'
        chart.series[1].title = 'Cat B'
        
        # Add to worksheet
        charts_ws.add_chart(chart, "B2")
        
        # Save workbook
        output_dir = root_dir / 'test_output'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / 'test_stacked_bar.xlsx'
        
        wb.save(str(output_path))
        logger.info(f"Test stacked bar chart saved to: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error in test_stacked_bar_only: {str(e)}")
        return False
    finally:
        wb.close()

def test_line_chart_only():
    """Test creating just a line chart."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add test data
        months = ['Jan-24', 'Feb-24']
        ws.cell(row=1, column=1, value='Series')  # Header for series column
        for col, month in enumerate(months, 2):
            ws.cell(row=1, column=col, value=month)
        
        # Add total row
        ws.cell(row=2, column=1, value='Total')
        ws.cell(row=2, column=2, value=150)
        ws.cell(row=2, column=3, value=180)
        
        # Log the data for debugging
        logger.info("Worksheet data:")
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
            logger.info(row)
        
        # Create Charts sheet
        charts_ws = wb.create_sheet('Charts')
        charts_ws.sheet_view.showGridLines = False
        
        # Create line chart
        chart = LineChart()
        chart.title = "Test Line Chart"
        chart.height = 15
        chart.width = 30
        chart.style = 10  # Try a specific style
        
        # Categories (X axis) - the months
        cats = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1)
        chart.set_categories(cats)
        
        # Data - total values
        data = Reference(ws, min_col=2, max_col=3, min_row=2, max_row=2)
        chart.add_data(data, titles_from_data=False)
        
        # Set series name
        chart.series[0].title = 'Total'
        
        # Add marker
        chart.series[0].marker = Marker(symbol='circle', size=10)
        chart.series[0].graphicalProperties.line.width = 2.25
        
        # Add to worksheet
        charts_ws.add_chart(chart, "B2")
        
        # Save workbook
        output_dir = root_dir / 'test_output'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / 'test_line.xlsx'
        
        wb.save(str(output_path))
        logger.info(f"Test line chart saved to: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error in test_line_chart_only: {str(e)}")
        return False
    finally:
        wb.close()

def test_stacked_bar_with_line():
    """Test creating a stacked bar chart with an overlaid line."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add test data
        months = ['Jan-24', 'Feb-24', 'Mar-24']
        for col, month in enumerate(months, 2):
            ws.cell(row=1, column=col, value=month)
        
        # Add some category data
        categories = ['Cat A', 'Cat B', 'Total']
        values = [
            [100, 120, 140],  # Cat A
            [50, 60, 70],     # Cat B
            [150, 180, 210]   # Total
        ]
        
        for row, (category, row_values) in enumerate(zip(categories, values), 2):
            ws.cell(row=row, column=1, value=category)
            for col, value in enumerate(row_values, 2):
                ws.cell(row=row, column=col, value=value)
        
        # Create Charts sheet
        charts_ws = wb.create_sheet('Charts')
        charts_ws.sheet_view.showGridLines = False
        
        # Create stacked bar chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Test Stacked Chart with Line"
        chart.height = 15
        chart.width = 30
        
        # Add category data (excluding total)
        cats = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=1)
        data = Reference(ws, min_col=2, min_row=2, max_col=4, max_row=3)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        # Add total line
        line = LineChart()
        line_data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=4)
        line.add_data(line_data, titles_from_data=False)
        line.y_axis.axId = 200
        line.y_axis.crosses = "max"
        
        # Add marker to line
        line.series[0].marker = Marker(symbol='circle')
        line.series[0].graphicalProperties.line.width = 2.25
        
        # Combine charts
        chart += line
        
        # Add to worksheet
        charts_ws.add_chart(chart, "B2")
        
        # Save workbook
        output_dir = root_dir / 'test_output'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / 'test_stacked_chart.xlsx'
        
        wb.save(str(output_path))
        logger.info(f"Test stacked chart saved to: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error in test_stacked_bar_with_line: {str(e)}")
        return False
    finally:
        wb.close()

def main():
    """Run chart type tests."""
    logger.info("\nTesting stacked bar chart only...")
    if test_stacked_bar_only():
        logger.info("✓ Stacked bar chart test passed")
    else:
        logger.error("✗ Stacked bar chart test failed")
        
    logger.info("\nTesting line chart only...")
    if test_line_chart_only():
        logger.info("✓ Line chart test passed")
    else:
        logger.error("✗ Line chart test failed")
        
    logger.info("\nTesting combined chart...")
    if test_stacked_bar_with_line():
        logger.info("✓ Combined chart test passed")
    else:
        logger.error("✗ Combined chart test failed")

if __name__ == "__main__":
    main() 